import os
import time
import hmac
import hashlib
import requests
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from pydantic import BaseModel
from fastapi import APIRouter, Request, HTTPException, UploadFile, File
from firebase_admin import auth as firebase_auth
from firebase_admin import firestore

from config.database import db
from config.limiter import limiter
from config.settings import ADMIN_EMAIL, ADMIN_EMAIL2
from middleware.auth import verify_admin, get_user_ratios
from utils.analytics import get_firebase_summary, get_dashboard_analytics

router = APIRouter()

class WithdrawalUpdate(BaseModel):
    request_id: str
    status: str

# Khởi tạo bộ nhớ đệm cho báo cáo Admin
ADMIN_REPORTS_CACHE = {
    "data": None,
    "last_updated": 0
}

@router.get("/api/admin/at-reports")
@limiter.limit("30/minute")
def admin_reports(request: Request):
    verify_admin(request)
    
    global ADMIN_REPORTS_CACHE
    import time
    now = time.time()
    
    # Trả về cache nếu chưa quá 10 phút (600 giây)
    if ADMIN_REPORTS_CACHE["data"] is not None and now - ADMIN_REPORTS_CACHE["last_updated"] < 600:
        return ADMIN_REPORTS_CACHE["data"]
        
    orders = [
        doc.to_dict()
        for doc in db.collection("orders").stream()
    ]
    firebase = get_firebase_summary()
    analytics = get_dashboard_analytics(orders)
    
    total_orders = len(orders)
    approved_commission = 0
    pending_commission = 0
    rejected_commission = 0

    approved_sales = 0
    pending_sales = 0
    rejected_sales = 0
    result = []
    
    approved_count = pending_count = reject_count = 0
    approved_admin_profit = 0
    
    for item in orders:
        commission = float(item.get("reward", 0))
        sales = float(item.get("product_price", 0))

        confirmed = int(item.get("confirmed", 0))
        status = int(item.get("status", 0))
        email = item.get("utm_source", "")

        if confirmed == 1:
            approved_commission += commission
            approved_sales += sales
            
            _, a_ratio, _ = get_user_ratios(email)
            approved_admin_profit += commission * a_ratio

            approved_count += 1
            order_status = 1
        elif status == 2:
            rejected_commission += commission
            rejected_sales += sales

            reject_count += 1
            order_status = 2
        else:
            pending_commission += commission
            pending_sales += sales

            pending_count += 1
            order_status = 0

        result.append({
            "order_id": item.get("order_id"),
            "order_time": item.get("sales_time"),
            "campaign_name": item.get("campaign_id"),
            "sales_amount": sales,
            "pub_commission": commission,
            "order_status": order_status,
            "utm_source": item.get("utm_source", "")
        })
        
    result.sort(key=lambda x: x["order_time"], reverse=True)
    
    report_data = {
        "success": True,
        "summary": {
            "conversions": total_orders,
            "approved_commission": round(approved_commission),
            "pending_commission": round(pending_commission),
            "rejected_commission": round(rejected_commission),
            "approved_sales": round(approved_sales),
            "pending_sales": round(pending_sales),
            "rejected_sales": round(rejected_sales),
            "net_profit": round(approved_admin_profit),
            "generated_links": firebase["generated_links"],
            "users": firebase["users"],
            "logs": firebase["logs"]
        },
        "analytics": {**analytics, "order_status": {"approved": approved_count, "pending": pending_count, "rejected": reject_count}},
        "orders": result
    }
    
    # Lưu kết quả vào Cache
    ADMIN_REPORTS_CACHE["data"] = report_data
    ADMIN_REPORTS_CACHE["last_updated"] = now
    
    return report_data

@router.get("/api/admin/withdrawals")
@limiter.limit("30/minute")
def get_withdrawals(request: Request):
    """Admin lấy danh sách yêu cầu rút tiền"""
    verify_admin(request)
    docs = db.collection("withdrawals").order_by("created_at", direction=firestore.Query.DESCENDING).stream()
    
    result = []
    for doc in docs:
        data = doc.to_dict()
        created_time = ""
        if "created_at" in data and data["created_at"]:
            try:
                created_time = data["created_at"].strftime("%d/%m/%Y %H:%M")
            except:
                created_time = str(data["created_at"])
        result.append({
            "id": doc.id,
            "email": data.get("user_email", "N/A"),
            "amount": data.get("amount", 0),
            "bank": data.get("bank_info", "N/A"),
            "status": data.get("status", "pending"),
            "date": created_time
        })
    return {"success": True, "data": result}

@router.post("/api/admin/withdrawals/update")
@limiter.limit("30/minute")
def update_withdrawal(request: Request, body: WithdrawalUpdate):
    """Admin cập nhật trạng thái yêu cầu rút tiền"""
    verify_admin(request)
    doc_ref = db.collection("withdrawals").document(body.request_id)
    
    if not doc_ref.get().exists:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu rút tiền này")
        
    doc_ref.update({
        "status": body.status,
        "updated_at": datetime.now()
    })
    
    # Xóa cache báo cáo để cập nhật lại số liệu mới
    global ADMIN_REPORTS_CACHE
    ADMIN_REPORTS_CACHE["last_updated"] = 0
    
    # Reset cả cache bảng xếp hạng bên user
    try:
        from routes.user import LEADERBOARD_CACHE
        LEADERBOARD_CACHE["last_updated"] = 0
    except:
        pass
        
    return {"success": True, "message": "Cập nhật trạng thái thành công"}

@router.get("/api/admin/users")
@limiter.limit("30/minute")
def get_admin_users(request: Request, start_date: str = None, end_date: str = None):
    """Admin lấy danh sách chi tiết hành vi người dùng có lọc theo ngày chuẩn xác"""
    verify_admin(request)
    
    query = db.collection("conversions")
    start_d = None
    end_d = None
    
    if start_date:
        try:
            start_d = datetime.strptime(start_date, "%Y-%m-%d").date()
            start_dt = datetime.combine(start_d, datetime.min.time(), tzinfo=timezone.utc)
            query = query.where("created_at", ">=", start_dt)
        except Exception:
            pass
    if end_date:
        try:
            end_d = datetime.strptime(end_date, "%Y-%m-%d").date()
            end_dt = datetime.combine(end_d + timedelta(days=1), datetime.min.time(), tzinfo=timezone.utc)
            query = query.where("created_at", "<", end_dt)
        except Exception:
            pass

    conversions = query.stream()
    user_data = defaultdict(lambda: {"email": "", "total_links": 0, "recent_links": []})

    for doc in conversions:
        data = doc.to_dict()
        email = data.get("user_email")
        if not email:
            continue
            
        created_at = data.get("created_at")
        time_str = "N/A"
        time_val = 0
        doc_d = None
        
        if created_at:
            try:
                if hasattr(created_at, "timestamp"):
                    time_val = created_at.timestamp()
                    dt_vn = datetime.fromtimestamp(time_val, tz=timezone.utc) + timedelta(hours=7)
                    doc_d = dt_vn.date()
                    time_str = dt_vn.strftime("%d/%m/%Y %H:%M")
                elif isinstance(created_at, str):
                    dt_obj = datetime.fromisoformat(created_at[:19])
                    doc_d = dt_obj.date()
                    time_val = dt_obj.timestamp()
                    time_str = dt_obj.strftime("%d/%m/%Y %H:%M")
            except Exception:
                pass

        if (start_d or end_d) and not doc_d:
            continue
            
        if doc_d:
            if start_d and doc_d < start_d:
                continue
            if end_d and doc_d > end_d:
                continue

        user_data[email]["email"] = email
        user_data[email]["total_links"] += 1
        
        user_data[email]["recent_links"].append({
            "product_name": data.get("product_name", "N/A"),
            "platform": data.get("platform", "N/A"),
            "short_link": data.get("short_link", "N/A"),
            "time_str": time_str,
            "time_val": time_val
        })
        
    result = []
    for email, info in user_data.items():
        info["recent_links"].sort(key=lambda x: x["time_val"], reverse=True)
        recent = info["recent_links"]
        if info["total_links"] > 0:
            result.append({
                "email": email,
                "total_links": info["total_links"],
                "recent_links": recent
            })
            
    result.sort(key=lambda x: x["total_links"], reverse=True)
    return {"success": True, "data": result}

@router.get("/api/admin/registered-users")
@limiter.limit("30/minute")
def get_registered_users(request: Request):
    """Admin lấy toàn bộ danh sách thành viên đăng ký kèm số dư ví"""
    verify_admin(request)
    
    # 1. Lấy toàn bộ danh sách users
    users_docs = db.collection("users").stream()
    users_list = []
    for doc in users_docs:
        users_list.append(doc.to_dict())
        
    # 2. Lấy toàn bộ đơn hàng để tính toán hoa hồng chéo (gộp theo email)
    orders_docs = db.collection("orders").stream()
    user_approved_cashback = defaultdict(float)
    user_pending_cashback = defaultdict(float)
    
    for doc in orders_docs:
        order = doc.to_dict()
        email = order.get("utm_source")
        if not email:
            continue
        confirmed = int(order.get("confirmed", 0))
        status = int(order.get("status", 0))
        reward = float(order.get("reward", 0))
        
        # Lấy tỷ lệ chia của user
        u_ratio, _, _ = get_user_ratios(email)
        cashback = reward * u_ratio
        
        if confirmed == 1:
            user_approved_cashback[email] += cashback
        elif status != 2:
            user_pending_cashback[email] += cashback
            
    # 3. Lấy toàn bộ withdrawals để tính toán rút tiền
    withdrawals_docs = db.collection("withdrawals").stream()
    user_approved_withdraw = defaultdict(float)
    user_pending_withdraw = defaultdict(float)
    
    for doc in withdrawals_docs:
        w = doc.to_dict()
        email = w.get("user_email")
        if not email:
            continue
        status = w.get("status")
        amount = float(w.get("amount", 0))
        
        if status == "approved":
            user_approved_withdraw[email] += amount
        elif status == "pending":
            user_pending_withdraw[email] += amount
            
    # 4. Gom dữ liệu trả về
    result = []
    for u in users_list:
        email = u.get("email", "")
        if not email:
            continue
            
        created_time = ""
        created_at_ms = u.get("createdAt")
        if created_at_ms:
            try:
                created_time = datetime.fromtimestamp(created_at_ms / 1000, tz=timezone.utc)
                created_time = (created_time + timedelta(hours=7)).strftime("%d/%m/%Y %H:%M")
            except:
                created_time = str(created_at_ms)
                
        # Tính toán khả dụng
        approved = user_approved_cashback.get(email, 0.0)
        app_withdraw = user_approved_withdraw.get(email, 0.0)
        pend_withdraw = user_pending_withdraw.get(email, 0.0)
        pending = user_pending_cashback.get(email, 0.0)
        
        available_balance = max(approved - app_withdraw - pend_withdraw, 0.0)
        
        result.append({
            "email": email,
            "displayName": u.get("displayName", ""),
            "photoURL": u.get("photoURL", ""),
            "createdAt": created_time,
            "provider": u.get("provider", ""),
            "balance": round(available_balance),
            "pending": round(pending),
            "withdrawn": round(app_withdraw)
        })
        
    result.sort(key=lambda x: x["createdAt"], reverse=True)
    return {"success": True, "data": result}

@router.post("/api/admin/sync-users")
@limiter.limit("5/minute")
def sync_users(request: Request):
    verify_admin(request)

    page = firebase_auth.list_users()
    total = 0

    while page:
        for user in page.users:
            email = user.email
            if not email:
                continue

            db.collection("users").document(email).set(
                {
                    "uid": user.uid,
                    "email": email,
                    "displayName": user.display_name or email.split("@")[0],
                    "photoURL": user.photo_url or "",
                    "phoneNumber": user.phone_number or "",
                    "disabled": user.disabled,
                    "createdAt": user.user_metadata.creation_timestamp,
                    "lastSignIn": user.user_metadata.last_sign_in_timestamp,
                    "provider": (
                        user.provider_data[0].provider_id
                        if user.provider_data
                        else ""
                    ),
                },
                merge=True,
            )
            total += 1
        page = page.get_next_page()

    return {
        "success": True,
        "synced": total
    }

@router.post("/api/admin/import-shopee-report")
@limiter.limit("5/minute")
async def import_shopee_report(request: Request, file: UploadFile = File(...)):
    verify_admin(request)
    
    try:
        contents = await file.read()
        filename = file.filename.lower()
        rows = []
        
        if filename.endswith(".csv"):
            # Try to decode CSV contents
            try:
                decoded = contents.decode("utf-8")
            except UnicodeDecodeError:
                try:
                    decoded = contents.decode("utf-16")
                except Exception as e:
                    raise Exception(f"Không thể giải mã file CSV: {e}")
            
            import csv
            # Strip UTF-8 BOM if present
            if decoded.startswith('\ufeff'):
                decoded = decoded[1:]
                
            lines = decoded.splitlines()
            if not lines:
                raise Exception("File CSV trống")
                
            # Detect delimiter (tab, semicolon, or comma)
            first_line = lines[0]
            delimiter = ","
            if "\t" in first_line:
                delimiter = "\t"
            elif ";" in first_line:
                delimiter = ";"
                
            reader = csv.reader(lines, delimiter=delimiter)
            rows = [row for row in reader if row]
        else:
            # Process as XLSX using openpyxl
            try:
                wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
                sheet = wb.active
                for r_idx in range(1, sheet.max_row + 1):
                    row_vals = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, sheet.max_column + 1)]
                    rows.append(row_vals)
            except Exception as e:
                raise Exception(f"Không thể đọc file XLSX: {str(e)}")
                
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không thể đọc file báo cáo: {str(e)}")
    
    header_row = None
    headers = []
    
    # Scan the first 15 rows to find the headers
    for r_idx, row in enumerate(rows[:15]):
        try:
            row_str_vals = [str(val or "").strip() for val in row]
            keywords = ["mã đơn hàng", "order id", "đơn hàng", "sub_id", "sub id", "hoa hồng", "commission"]
            if any(any(kw in val.lower() for kw in keywords) for val in row_str_vals):
                header_row = r_idx
                headers = row_str_vals
                break
        except Exception:
            continue
            
    if header_row is None:
        raise HTTPException(status_code=400, detail="Không tìm thấy dòng tiêu đề hợp lệ trong file báo cáo. Vui lòng kiểm tra lại file báo cáo của Shopee.")
        
    col_map = {"sub_ids": []}
    for idx, name in enumerate(headers):
        name_lower = name.lower().strip()
        
        # Match order_id (ID đơn hàng / Mã đơn hàng)
        if any(x in name_lower for x in ["mã đơn hàng", "ma don hang", "order id", "order_id", "id đơn hàng", "id don hang", "id đơn", "id don", "mã đơn", "ma don"]):
            col_map["order_id"] = idx
            
        # Match status (Trạng thái)
        elif any(x in name_lower for x in ["trạng thái", "trang thai", "status"]):
            if not any(x in name_lower for x in ["người mua", "nguoi mua", "tài khoản", "tai khoan"]):
                col_map["status"] = idx
            
        # Match product_price (Giá trị đơn hàng)
        elif any(x in name_lower for x in ["giá trị đơn", "gia tri don", "order value", "doanh số", "doanh so", "giá trị sản phẩm", "gia tri san pham", "product price", "price", "giá bán", "gia ban", "giá trị đơn hàng", "gia tri don hang"]):
            col_map["product_price"] = idx
            
        # Match reward (Hoa hồng) - Exclude rate/type/level columns
        elif any(x in name_lower for x in ["hoa hồng", "hoa hong", "commission", "reward"]):
            if not any(x in name_lower for x in ["tỉ lệ", "ti le", "loại", "loai", "mức", "muc", "tỷ lệ", "ty le"]):
                col_map["reward"] = idx
                
        # Match sub_id (sub_id1, sub_id2, sub_id, utm_source)
        elif any(x in name_lower for x in ["sub_id", "sub id", "sub-id", "subid", "utm_content", "sub_id1", "sub_id2", "sub_id3", "sub_id4", "sub_id5", "sub_id 1", "sub_id 2", "utm_source", "utm source"]):
            col_map["sub_ids"].append(idx)
            
        # Match sales_time (Thời gian đặt hàng)
        elif any(x in name_lower for x in ["thời gian đặt", "thoi gian dat", "sales time", "sales_time", "order time", "thời gian tạo", "thoi gian tao", "thời gian mua", "thoi gian mua"]):
            col_map["sales_time"] = idx
        elif "thời gian" in name_lower or "thoi gian" in name_lower or "time" in name_lower:
            if "sales_time" not in col_map:
                col_map["sales_time"] = idx
                
        # Match transaction_id
        elif any(x in name_lower for x in ["mã lượt click", "ma luot click", "click id", "transaction_id", "transaction id"]):
            col_map["transaction_id"] = idx
            
        # Match advertiser/campaign
        elif any(x in name_lower for x in ["advertiser", "campaign", "chiến dịch", "chien dich", "nhà quảng cáo", "nha quang cao"]):
            col_map["advertiser"] = idx

    required_cols = ["order_id", "status", "reward"]
    missing = [c for c in required_cols if c not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"File báo cáo thiếu các cột bắt buộc: {', '.join(missing)}")

    success_count = 0
    skipped_count = 0
    
    # Process from header_row + 1 onwards
    for r_idx, row_vals in enumerate(rows[header_row + 1:]):
        try:
            if not row_vals or all(val is None or str(val).strip() == "" for val in row_vals):
                continue
                
            # Sửa lỗi tách cột do dấu phẩy ở cột Mã danh mục ("bonus,xxxxx") của AccessTrade
            cat_header_idx = None
            for idx, h in enumerate(headers):
                if "danh mục" in h.lower() or "danh muc" in h.lower():
                    cat_header_idx = idx
                    break
            if cat_header_idx is not None and len(row_vals) > len(headers):
                diff = len(row_vals) - len(headers)
                merged_val = ",".join(str(row_vals[i]) if row_vals[i] is not None else "" for i in range(cat_header_idx, cat_header_idx + diff + 1))
                row_vals = row_vals[:cat_header_idx] + [merged_val] + row_vals[cat_header_idx + diff + 1:]

            def get_val(col_name, default=None):
                idx = col_map.get(col_name)
                if idx is not None and idx < len(row_vals):
                    return row_vals[idx]
                return default

            raw_order_id = get_val("order_id")
            if not raw_order_id:
                skipped_count += 1
                continue
                
            order_id = str(raw_order_id).strip()
            raw_status = get_val("status", "")
            status_str = str(raw_status).strip().lower()
            
            confirmed = 0
            status_code = 0
            
            if "pre approved" in status_str or "pre_approved" in status_str or "tạm" in status_str or "chờ" in status_str:
                confirmed = 0
                status_code = 0
            elif any(x in status_str for x in ["hoàn thành", "thành công", "completed", "đã hoàn thành", "được duyệt", "duyệt", "approved"]):
                confirmed = 1
                status_code = 1
            elif any(x in status_str for x in ["hủy", "cancelled", "không thành công", "đã hủy", "bị hủy", "rejected", "reject"]):
                confirmed = 0
                status_code = 2
            else:
                confirmed = 0
                status_code = 0
                
            def parse_float(val):
                if val is None:
                    return 0.0
                try:
                    s = str(val).replace(",", "").replace("đ", "").replace("VND", "").strip()
                    return float(s)
                except:
                    return 0.0

            reward = parse_float(get_val("reward"))
            product_price = parse_float(get_val("product_price"))
            
            # Dynamic Advertiser / Campaign detection
            raw_campaign = get_val("advertiser", "Shopee")
            campaign_id = str(raw_campaign).strip()
            camp_lower = campaign_id.lower()
            
            if "tiktok" in camp_lower:
                utm_medium = "tiktok"
                # Keep original or map to TikTok Shop
                if campaign_id == "Shopee" or not campaign_id:
                    campaign_id = "TikTok Shop"
            elif "lazada" in camp_lower:
                utm_medium = "lazada"
                if campaign_id == "Shopee" or not campaign_id:
                    campaign_id = "Lazada"
            elif "shopee" in camp_lower:
                utm_medium = "shopee"
            else:
                utm_medium = "shopee"

            # Extract sub_ids values and search for email containing _at_
            sub_id_vals = []
            for sub_idx in col_map.get("sub_ids", []):
                if sub_idx < len(row_vals) and row_vals[sub_idx]:
                    sub_id_vals.append(str(row_vals[sub_idx]).strip())
            
            email = ""
            for val in sub_id_vals:
                if "_at_" in val:
                    sanitized_email = val
                    parts = sanitized_email.split("_at_")
                    username = parts[0]
                    domain = parts[1].replace("_", ".")
                    
                    if not domain.endswith("com") and not domain.endswith("vn") and not domain.endswith("net"):
                        if domain.startswith("gmail"):
                            domain = "gmail.com"
                        elif domain.startswith("yahoo"):
                            domain = "yahoo.com"
                    temp_email = f"{username}@{domain}"
                    
                    matched_user_email = ""
                    try:
                        query = db.collection("users").where("email", ">=", username).where("email", "<", username + "\uf8ff").limit(5).stream()
                        for doc in query:
                            u_email = doc.to_dict().get("email", "")
                            u_sanitized = u_email.replace("-", "_").replace("@", "_at_").replace(".", "_")
                            if u_sanitized.startswith(sanitized_email) or sanitized_email.startswith(u_sanitized[:len(sanitized_email)]):
                                matched_user_email = u_email
                                break
                    except Exception as ex:
                        print(f"Firestore sub_id matching error: {ex}")
                        
                    email = matched_user_email if matched_user_email else temp_email
                    break
            else:
                # Fallback to direct email checking in any sub_id field (like utm_source in TikTok/AccessTrade)
                for val in sub_id_vals:
                    if "@" in val:
                        email = val
                        break

            # Reconstruct sub_id for utm_content logging
            sub_id = "-".join(sub_id_vals)
            
            raw_tx_id = get_val("transaction_id")
            if raw_tx_id:
                transaction_id = str(raw_tx_id).strip()
            else:
                transaction_id = f"{utm_medium}_{order_id}"

            sales_time = get_val("sales_time")
            if sales_time:
                if hasattr(sales_time, "strftime"):
                    sales_time_str = sales_time.strftime("%Y-%m-%dT%H:%M:%S")
                else:
                    sales_time_str = str(sales_time).strip().replace(" ", "T")
            else:
                sales_time_str = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

            db.collection("orders").document(transaction_id).set({
                "transaction_id": transaction_id,
                "order_id": order_id,
                "campaign_id": campaign_id,
                "product_id": "",
                "quantity": 1,
                "product_price": product_price,
                "reward": reward,
                "sales_time": sales_time_str,
                "status": status_code,
                "confirmed": confirmed,
                "utm_source": email,
                "utm_content": sub_id,
                "utm_medium": utm_medium,
                "created_at": firestore.SERVER_TIMESTAMP
            }, merge=True)
            
            success_count += 1
        except Exception as e:
            print(f"Lỗi khi đọc dòng {r_idx}: {e}")
            skipped_count += 1
            
    # Reset cache báo cáo Admin và bảng xếp hạng khi import thành công
    global ADMIN_REPORTS_CACHE
    ADMIN_REPORTS_CACHE["last_updated"] = 0
    try:
        from routes.user import LEADERBOARD_CACHE
        LEADERBOARD_CACHE["last_updated"] = 0
    except:
        pass
        
    return {
        "success": True,
        "message": f"Đã nhập thành công {success_count} đơn hàng Shopee. Bỏ qua {skipped_count} dòng lỗi."
    }

# --- LAZADA INTEGRATION API FOR SYNCING ORDERS ---
LAZADA_APP_KEY = os.getenv("LAZADA_APP_KEY")
LAZADA_APP_SECRET = os.getenv("LAZADA_APP_SECRET")
LAZADA_USER_TOKEN = os.getenv("LAZADA_USER_TOKEN")

def generate_lazada_sign(api_path: str, params: dict, app_secret: str) -> str:
    """Tính toán chữ ký HMAC-SHA256 theo tiêu chuẩn của Lazada Open Platform"""
    sorted_params = sorted(params.items())
    query_str = api_path
    for key, val in sorted_params:
        query_str += f"{key}{val}"
    sign = hmac.new(
        app_secret.encode("utf-8"),
        query_str.encode("utf-8"),
        hashlib.sha256
    ).hexdigest().upper()
    return sign

@router.post("/api/admin/sync-lazada")
@limiter.limit("5/minute")
def sync_lazada_orders(request: Request, days: int = 7):
    verify_admin(request)
    
    if not LAZADA_APP_KEY or not LAZADA_APP_SECRET or not LAZADA_USER_TOKEN:
        raise HTTPException(
            status_code=500,
            detail="Chưa cấu hình các biến môi trường LAZADA_APP_KEY, LAZADA_APP_SECRET, hoặc LAZADA_USER_TOKEN trên máy chủ."
        )

    # 1. Tính toán khoảng thời gian đồng bộ (Ví dụ: 7 ngày qua)
    # Định dạng ngày: YYYY-MM-DD
    # Múi giờ Việt Nam (UTC+7)
    vn_tz = timezone(timedelta(hours=7))
    now_vn = datetime.now(vn_tz)
    start_date = (now_vn - timedelta(days=days)).strftime("%Y-%m-%d")
    end_date = now_vn.strftime("%Y-%m-%d")

    base_url = "https://api.lazada.vn/rest"
    api_path = "/marketing/conversion/report"

    # API hỗ trợ phân trang, ta chạy tối đa 10 trang (tối đa 1000 đơn hàng) để tránh treo server
    imported_count = 0
    updated_count = 0
    page = 1
    limit = 100

    while page <= 10:
        params = {
            "app_key": LAZADA_APP_KEY,
            "timestamp": str(int(time.time() * 1000)),
            "sign_method": "sha256",
            "userToken": LAZADA_USER_TOKEN,
            "dateStart": start_date,
            "dateEnd": end_date,
            "page": str(page),
            "limit": str(limit)
        }
        params["sign"] = generate_lazada_sign(api_path, params, LAZADA_APP_SECRET)

        try:
            res = requests.get(f"{base_url}{api_path}", params=params, timeout=15)
            if res.status_code != 200:
                print(f"Lazada API sync failed page {page}: HTTP {res.status_code}")
                break
                
            res_data = res.json()
            result = res_data.get("result", {})
            success = result.get("success")
            orders_list = result.get("data", [])
            
            if not success or not orders_list:
                break

            for order_item in orders_list:
                order_id = str(order_item.get("orderId"))
                sub_order_id = str(order_item.get("subOrderId"))
                transaction_id = f"lazada_{sub_order_id}"
                
                # Trạng thái đơn hàng
                status_str = str(order_item.get("status", "")).lower()
                
                # Quy đổi trạng thái đơn hàng
                # status_code: 0 = pending, 1 = approved, 2 = rejected
                if "delivered" in status_str:
                    confirmed = 1
                    status_code = 1
                elif any(x in status_str for x in ["return", "cancel", "refund", "reject"]):
                    confirmed = 0
                    status_code = 2
                else:
                    confirmed = 0
                    status_code = 0

                # Số tiền đơn hàng và hoa hồng
                product_price = float(order_item.get("orderAmt", 0.0))
                reward = float(order_item.get("estPayout", 0.0))
                
                # Giải mã email khách hàng từ subId1
                sub_id1 = order_item.get("subId1") or ""
                email = ""
                
                if sub_id1:
                    if "_at_" in sub_id1:
                        parts = sub_id1.split("_at_")
                        username = parts[0]
                        domain = parts[1].replace("_", ".")
                        # Chuẩn hóa tên miền
                        if not domain.endswith("com") and not domain.endswith("vn") and not domain.endswith("net"):
                            if domain.startswith("gmail"):
                                domain = "gmail.com"
                            elif domain.startswith("yahoo"):
                                domain = "yahoo.com"
                        temp_email = f"{username}@{domain}"
                        
                        # Khớp chéo database để lấy email thật chính xác nếu bị cắt ngắn
                        matched_user_email = ""
                        try:
                            query = db.collection("users").where("email", ">=", username).where("email", "<", username + "\uf8ff").limit(3).stream()
                            for doc in query:
                                u_email = doc.to_dict().get("email", "")
                                u_sanitized = u_email.replace("-", "_").replace("@", "_at_").replace(".", "_")
                                if u_sanitized.startswith(sub_id1) or sub_id1.startswith(u_sanitized[:len(sub_id1)]):
                                    matched_user_email = u_email
                                    break
                        except Exception:
                            pass
                        email = matched_user_email if matched_user_email else temp_email
                    elif "@" in sub_id1:
                        email = sub_id1

                # Thời gian tạo đơn
                conversion_time = order_item.get("conversionTime")
                if conversion_time:
                    sales_time_str = str(conversion_time).replace(" ", "T")
                else:
                    sales_time_str = datetime.now(vn_tz).strftime("%Y-%m-%dT%H:%M:%S")

                # Lưu vào Firestore
                order_doc = db.collection("orders").document(transaction_id)
                existing_doc = order_doc.get()
                
                order_payload = {
                    "transaction_id": transaction_id,
                    "order_id": order_id,
                    "campaign_id": "Lazada",
                    "product_id": str(order_item.get("sku", "")),
                    "quantity": 1,
                    "product_price": product_price,
                    "reward": reward,
                    "sales_time": sales_time_str,
                    "status": status_code,
                    "confirmed": confirmed,
                    "utm_source": email,
                    "utm_content": sub_id1,
                    "utm_medium": "lazada",
                    "created_at": firestore.SERVER_TIMESTAMP
                }
                
                order_doc.set(order_payload, merge=True)
                
                if existing_doc.exists:
                    updated_count += 1
                else:
                    imported_count += 1
                    
            if len(orders_list) < limit:
                break
            page += 1
        except Exception as e:
            print(f"Lỗi khi đồng bộ trang {page}: {e}")
            break

    # Reset cache báo cáo Admin và bảng xếp hạng khi đồng bộ thành công
    global ADMIN_REPORTS_CACHE
    ADMIN_REPORTS_CACHE["last_updated"] = 0
    try:
        from routes.user import LEADERBOARD_CACHE
        LEADERBOARD_CACHE["last_updated"] = 0
    except:
        pass

    return {
        "success": True,
        "message": f"Đồng bộ đơn hàng Lazada hoàn tất từ ngày {start_date} đến {end_date}.",
        "imported": imported_count,
        "updated": updated_count
    }

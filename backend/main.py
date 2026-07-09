import os
import requests
import json
import random
import string
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime, timedelta, timezone
from firebase_admin import auth as firebase_auth
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from fastapi import Request
from fastapi.responses import RedirectResponse
from slowapi import _rate_limit_exceeded_handler
from collections import Counter, defaultdict
from urllib.parse import quote, urlparse, urlunparse

# 1. Load cấu hình
load_dotenv()
SHOPEE_AFFILIATE_ID = os.getenv("SHOPEE_AFFILIATE_ID")
ENABLE_SHOPEE = os.getenv("ENABLE_SHOPEE", "false").lower() == "true"
AT_API_KEY = os.getenv("ACCESSTRADE_API_KEY")

# Campaign IDs
TIKTOK_CAMPAIGN_ID = os.getenv("TIKTOK_CAMPAIGN_ID")
SHOPEE_CAMPAIGN_ID = os.getenv("SHOPEE_CAMPAIGN_ID")
LAZADA_CAMPAIGN_ID = os.getenv("LAZADA_CAMPAIGN_ID")

# Tài khoản admin
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL")
ADMIN_EMAIL2 = os.getenv("ADMIN_EMAIL2")
cashback_percent = 70
user_ratio = 0.7
admin_ratio = 0.3
REPORT_DAYS = 30
REQUEST_TIMEOUT = 15

def get_user_ratios(email: str):
    """
    Trả về (user_ratio, admin_ratio, cashback_percent) động theo email.
    Đối với daonam5696@gmail.com được hưởng trọn 100% hoa hồng.
    """
    if email and email.strip().lower() == "daonam5696@gmail.com":
        return 1.0, 0.0, 100
    return user_ratio, admin_ratio, cashback_percent

# 2. Khởi tạo Firebase Admin
try:
    firebase_creds_json = os.getenv("FIREBASE_CREDENTIALS")
    if firebase_creds_json:
        creds_dict = json.loads(firebase_creds_json)
        cred = credentials.Certificate(creds_dict)
        firebase_admin.initialize_app(cred)
        db = firestore.client()
        print("Firebase connected successfully via Env Var!")
    else:
        cred = credentials.Certificate(
            os.path.join(os.path.dirname(__file__), "firebase-key.json")
        )
        firebase_admin.initialize_app(cred)
        db = firestore.client()
        print("Firebase connected successfully via Local File!")
except Exception as e:
    print(f"Firebase initialization error: {e}")

# 3. Khởi tạo FastAPI
app = FastAPI()
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(SlowAPIMiddleware)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "https://tho-hoan-tien.vercel.app",
        "https://www.tho-hoantien.com",
        "https://tho-hoantien.com",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 4. Models
class LinkRequest(BaseModel):
    user_email: str
    original_url: str
    platform: str

class WithdrawalRequest(BaseModel):
    user_email: str
    amount: float
    bank_info: str

class WithdrawalUpdate(BaseModel):
    request_id: str
    status: str # "approved" hoặc "rejected"

# ==========================================
# CÁC HÀM TIỆN ÍCH (UTILS)
# ==========================================
def generate_short_code(length=6):
    chars = string.ascii_letters + string.digits
    return "".join(random.choice(chars) for _ in range(length))

def clean_shopee_url(url: str) -> str:
    try:
        parsed = urlparse(url)
        if "shopee.vn" in parsed.netloc:
            return urlunparse((parsed.scheme, parsed.netloc, parsed.path, '', '', ''))
    except Exception:
        pass
    return url

# ==========================================
# ENDPOINT: NHẬN WEBHOOK TỪ ACCESSTRADE
# ==========================================
@app.api_route("/api/postback", methods=["GET", "POST"])
async def global_postback(request: Request):
    p = dict(request.query_params)
    if request.method == "POST":
        try:
            body_json = await request.json()
            p.update(body_json)
        except Exception:
            pass

    def safe_float(val):
        try: return float(val) if val else 0.0
        except ValueError: return 0.0
        
    def safe_int(val):
        try: return int(val) if val else 0
        except ValueError: return 0

    # --- ACCESSTRADE WEBHOOK ---
    transaction_id = p.get("transaction_id")
    if not transaction_id:
        return {"success": True, "message": "AccessTrade ping ok"}

    sales_time_str = str(p.get("sales_time", ""))
    if sales_time_str:
        sales_time_str = sales_time_str.replace(" ", "T")

    db.collection("orders").document(str(transaction_id)).set({
        "transaction_id": str(transaction_id),
        "order_id": str(p.get("order_id", "")),
        "campaign_id": str(p.get("campaign_id", "")),
        "product_id": str(p.get("product_id", "")),
        "quantity": safe_int(p.get("quantity")),
        "product_price": safe_float(p.get("product_price")),
        "reward": safe_float(p.get("reward")),
        "sales_time": sales_time_str,
        "status": safe_int(p.get("status")),
        "confirmed": safe_int(p.get("is_confirmed")),
        "utm_source": str(p.get("utm_source", "")),
        "utm_campaign": str(p.get("utm_campaign", "")),
        "utm_medium": str(p.get("utm_medium", "")),
        "utm_content": str(p.get("utm_content", "")),
        "browser": str(p.get("browser", "")),
        "platform": str(p.get("conversion_platform", "")),
        "ip": str(p.get("ip", "")),
        "created_at": firestore.SERVER_TIMESTAMP,
        "raw": p
    }, merge=True)

    return {"success": True}

@app.get("/r/{code}")
def redirect_short_url(code: str):
    doc_ref = db.collection("short_urls").document(code).get()
    if not doc_ref.exists:
        raise HTTPException(status_code=404, detail="Đường liên kết không tồn tại hoặc đã hết hạn.")
    data = doc_ref.to_dict()
    long_url = data.get("long_url")
    return RedirectResponse(url=long_url, status_code=307)

def verify_admin(request: Request):
    auth_header = request.headers.get("Authorization")
    if not auth_header:
        raise HTTPException(status_code=401, detail="Missing token")
    token = auth_header.replace("Bearer ", "")
    try:
        decoded = firebase_auth.verify_id_token(token)
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")
    email = decoded.get("email")
    if email != ADMIN_EMAIL and email != ADMIN_EMAIL2:
        raise HTTPException(status_code=403, detail="Forbidden")
    return decoded

def get_firebase_summary():
    conversions = list(db.collection("conversions").stream())
    logs = list(db.collection("logs").stream())
    emails = set()
    for doc in conversions:
        data = doc.to_dict()
        email = data.get("user_email")
        if email:
            emails.add(email)
    return {
        "generated_links": len(conversions),
        "logs": len(logs),
        "users": len(emails)
    }

def get_dashboard_analytics(orders):
    conversions = [doc.to_dict() for doc in db.collection("conversions").stream()]
    today = datetime.now(timezone.utc) + timedelta(hours=7)
    today_date = today.date()
    week_ago = today_date - timedelta(days=6)
    
    daily_links = defaultdict(int)
    first_seen = {}
    product_counter = Counter()
    user_cashback = defaultdict(float)
    today_links = 0
    
    for item in conversions:
        created = item.get("created_at")
        doc_date = None
        if created:
            if hasattr(created, "timestamp"):
                doc_date = (datetime.fromtimestamp(created.timestamp(), tz=timezone.utc) + timedelta(hours=7)).date()
            elif isinstance(created, str):
                try: doc_date = datetime.fromisoformat(created[:19]).date()
                except: pass
                
        if not doc_date: continue
        if doc_date >= week_ago: daily_links[str(doc_date)] += 1
        if doc_date == today_date: today_links += 1
            
        email = item.get("user_email")
        if email and (email not in first_seen or doc_date < first_seen[email]):
            first_seen[email] = doc_date
            
        product = item.get("product_name")
        platform = item.get("platform", "tiktok")
        if product: product_counter[(product, platform)] += 1

    for order in orders:
        email = order.get("utm_source")
        if not email:
            continue

        confirmed = int(order.get("confirmed", 0))

        if confirmed != 1:
            continue

        reward = float(order.get("reward", 0))
        u_ratio, _, _ = get_user_ratios(email)
        user_cashback[email] += reward * u_ratio
    
    new_users = sum(1 for d in first_seen.values() if d >= week_ago)
    chart = [{"date": (week_ago + timedelta(days=i)).strftime("%d/%m"), "count": daily_links.get(str(week_ago + timedelta(days=i)), 0)} for i in range(7)]
        
    return {
        "daily_links": chart,
        "top_users": [{"email": k, "cashback": round(v)} for k, v in sorted(user_cashback.items(), key=lambda x: x[1], reverse=True)[:10]],
        "top_products": [{"name": k[0], "platform": k[1], "count": v} for k, v in product_counter.most_common(10)],
        "new_users": new_users,
        "today_links": today_links
    }


# ==========================================
# ENDPOINT: LẤY THÔNG TIN VÍ
# ==========================================
# GHI CHÚ QUAN TRỌNG: Hàm này đã được thêm `request: Request` để không bị lỗi SlowAPI.
@app.get("/api/user/wallet")
@limiter.limit("30/minute")
def get_user_wallet(email: str, request: Request):

    orders = db.collection("orders") \
        .where("utm_source", "==", email) \
        .stream()

    total_approved = 0
    total_pending = 0
    for doc in orders:
        order = doc.to_dict()
        confirmed = int(order.get("confirmed", 0))
        status = int(order.get("status", 0))
        u_ratio, _, _ = get_user_ratios(email)
        cashback = float(order.get("reward", 0)) * u_ratio

        print("==============")
        print(order)
        print("confirmed =", confirmed)
        print("status =", status)
        print("cashback =", cashback)

        if confirmed == 1:
            print("=> APPROVED")
            total_approved += cashback
        elif status != 2:
            print("=> PENDING")
            total_pending += cashback

    print("approved =", total_approved)
    print("pending =", total_pending)

    approved_withdraw = sum(
        d.to_dict().get("amount", 0)
        for d in db.collection("withdrawals")
        .where("user_email", "==", email)
        .where("status", "==", "approved")
        .stream()
    )

    pending_withdraw = sum(
        d.to_dict().get("amount", 0)
        for d in db.collection("withdrawals")
        .where("user_email", "==", email)
        .where("status", "==", "pending")
        .stream()
    )

    available = max(
        total_approved - approved_withdraw - pending_withdraw,
        0
    )
    print([doc.to_dict() for doc in db.collection("orders").where("utm_source", "==", email).stream()])
    return {
        "success": True,
        "balance": round(available),
        "pending": round(total_pending),
        "withdrawn": round(approved_withdraw)
    }


# ==========================================
# ENDPOINT: RÚT GỌN LINK
# ==========================================
@app.post("/api/convert")
@limiter.limit("30/minute") 
async def convert_link(request: Request, body: LinkRequest):
    headers = {
        "Authorization": f"Token {AT_API_KEY}",
        "Content-Type": "application/json",
        "accept": "application/json"
    }

    if body.platform == "tiktok":
        payload = {
            "product_url": body.original_url,
            "utm_source": body.user_email,
            "utm_medium": body.platform,
            "utm_campaign": "cashback"
        }
        response = requests.post(
            "https://api.accesstrade.vn/v2/tiktokshop_product_feeds/create_link",
            headers=headers,
            json=payload,
            timeout=REQUEST_TIMEOUT
        )
        if response.status_code != 200:
            raise HTTPException(status_code=500, detail="Không thể kết nối AccessTrade")
        response_data = response.json()
        if not response_data.get("status"):
            msg = response_data.get("message", "Không thể tạo link")
            if msg == "invalid params": msg = "Link không hợp lệ hoặc chưa hỗ trợ hoàn tiền."
            elif "campaign" in msg.lower(): msg = "Sản phẩm chưa tham gia hoàn tiền."
            elif "not found" in msg.lower(): msg = "Không tìm thấy sản phẩm."
            elif "policy" in msg.lower() or "failed to be processed" in msg.lower(): 
                msg = "Sản phẩm này không hỗ trợ hoàn tiền đâu nè (do vi phạm chính sách sản phẩm của sàn) 🐰"
            raise HTTPException(status_code=400, detail=msg)
        data = response_data["data"]
        aff_link = data["aff_url"]
        short_link = data["aff_short_url"]
        product_name = data["product_name"]
        product_image = data["product_image"]
        
        price_info = data.get("product_price", {})
        product_price = float(price_info.get("maximum_amount") or price_info.get("minimum_amount") or 0)
        
        commission_info = data.get("product_commission", {})
        commission = float(commission_info.get("amount", 0))
        u_ratio, a_ratio, c_percent = get_user_ratios(body.user_email)
        cashback = round(commission * u_ratio)
        publisher_income = round(commission * a_ratio)
    elif body.platform == "shopee":
        if not ENABLE_SHOPEE:
            raise HTTPException(
                status_code=400,
                detail="Hoàn tiền Shopee đang được chuẩn bị và sẽ ra mắt sớm! Hiện tại bạn hãy trải nghiệm mua sắm qua TikTok Shop nhé 🐰"
            )
        product_name = f"Sản phẩm Shopee"
        product_image = ""
        product_price = 0.0
        commission = 0.0

        # 1. Làm sạch link sản phẩm gốc trước khi gửi đi
        cleaned_url = clean_shopee_url(body.original_url)

        try:
            data_api_url = f"https://data.addlivetag.com/product-data/product-data.php?url={quote(cleaned_url)}"
            response_data = requests.get(data_api_url, timeout=REQUEST_TIMEOUT)
            if response_data.status_code == 200:
                res_json = response_data.json()
                if res_json.get("status") == "success" and res_json.get("productInfo"):
                    p_info = res_json["productInfo"]
                    product_name = p_info.get("productName", product_name)
                    product_price = float(p_info.get("price", 0.0))
                    product_image = p_info.get("imageUrl", "")
                    commission = float(p_info.get("commission", 0.0))
        except Exception as e:
            print(f"Shopee Product Data API error: {e}")

        if not product_image:
            product_image = "https://upload.wikimedia.org/wikipedia/commons/thumb/f/fe/Shopee.svg/375px-Shopee.svg.png"

        # 2. Tạo link an_redir trực tiếp bằng link đã làm sạch
        sanitized_email = body.user_email.replace("-", "_").replace("@", "_at_").replace(".", "_")
        sub_id = f"hangthocashback-{sanitized_email}"
        encoded_url = quote(cleaned_url, safe="")
        aff_link = f"https://s.shopee.vn/an_redir?origin_link={encoded_url}&affiliate_id={SHOPEE_AFFILIATE_ID}&sub_id={sub_id}"

        # 3. Tạo link ngắn custom của hệ thống ta qua endpoint /r/{code}
        short_code = generate_short_code()
        db.collection("short_urls").document(short_code).set({
            "long_url": aff_link,
            "created_at": firestore.SERVER_TIMESTAMP
        })
        base_url = str(request.base_url).rstrip("/")
        short_link = f"{base_url}/r/{short_code}"

        u_ratio, a_ratio, c_percent = get_user_ratios(body.user_email)
        cashback = round(commission * u_ratio)
        publisher_income = round(commission * a_ratio)

    elif body.platform == "lazada":
        campaign_id = LAZADA_CAMPAIGN_ID
        if not campaign_id:
            raise HTTPException(
                status_code=400, 
                detail="Hoàn tiền Lazada đang được chuẩn bị và sẽ ra mắt sớm! Hiện tại bạn hãy trải nghiệm mua sắm qua TikTok Shop nhé 🐰"
            )

        payload = {
            "campaign_id": campaign_id,
            "urls": [body.original_url],
            "utm_source": body.user_email,
            "utm_medium": body.platform,
            "utm_campaign": "cashback"
        }
        response = requests.post(
            "https://api.accesstrade.vn/v1/product_link/create",
            headers=headers,
            json=payload,
            timeout=REQUEST_TIMEOUT
        )
        if response.status_code != 200:
            raise HTTPException(status_code=500, detail=f"Không thể kết nối AccessTrade ({response.status_code})")
        
        response_data = response.json()
        if not response_data.get("success"):
            raise HTTPException(status_code=400, detail="Không thể tạo link liên kết cho sản phẩm này.")
            
        success_links = response_data.get("data", {}).get("success_link", [])
        if not success_links:
            raise HTTPException(status_code=400, detail="Tạo link thất bại. Vui lòng kiểm tra lại liên kết sản phẩm.")
            
        link_data = success_links[0]
        aff_link = link_data["aff_link"]
        short_link = link_data["short_link"]
        
        # Trích xuất tên sản phẩm từ URL
        product_name = f"Sản phẩm Lazada"
        if "lazada.vn/products/" in body.original_url:
            try:
                parts = body.original_url.split("lazada.vn/products/")[1].split("?")[0].split(".html")[0].split("-")
                if len(parts) > 0:
                    product_name = " ".join(parts[:-1])
            except Exception:
                pass
                
        product_image = "https://upload.wikimedia.org/wikipedia/commons/0/06/Lazada_Logo.png"
            
        product_price = 0.0
        commission = 0.0
        cashback = 0.0
        u_ratio, a_ratio, c_percent = get_user_ratios(body.user_email)
        publisher_income = 0.0
    else:
        raise HTTPException(status_code=400, detail="Nền tảng không hợp lệ")

    client_ip = request.client.host

    db.collection("logs").add({
        "ip": client_ip, "email": body.user_email, "platform": body.platform,
        "url": body.original_url, "product_name": product_name, "created_at": firestore.SERVER_TIMESTAMP
    })
    db.collection("conversions").add({
        "user_email": body.user_email,
        "original_url": body.original_url,
        "platform": body.platform,
        "product_name": product_name,
        "product_price": product_price,
        "short_link": short_link,
        "aff_link": aff_link,
        "created_at": firestore.SERVER_TIMESTAMP,
        "status": "link_created"
    })
    
    # Lấy lại c_percent động cho email
    _, _, c_percent = get_user_ratios(body.user_email)
    
    return {
        "success": True,
        "product": {"name": product_name, "image": product_image, "price": product_price},
        "commission": {"amount": commission, "cashback_percent": c_percent, "cashback": cashback, "publisher_income": publisher_income},
        "links": {"short": short_link, "affiliate": aff_link}
    }

# ==========================================
# ENDPOINTS: ADMIN BÁO CÁO
# ==========================================
@app.get("/api/admin/at-reports")
@limiter.limit("30/minute")
def admin_reports(request: Request):
    verify_admin(request)
    orders = [
        doc.to_dict()
        for doc in db.collection("orders").stream()
    ]
    firebase = get_firebase_summary()
    analytics = get_dashboard_analytics(orders)
    
    total_orders = len(orders)
    approved_commission = 0
    pending_commission = 0
    rejected_commission = 0

    approved_sales = 0
    pending_sales = 0
    rejected_sales = 0
    result = []
    
    approved_count = pending_count = reject_count = 0
    approved_admin_profit = 0
    
    for item in orders:

        commission = float(item.get("reward", 0))
        sales = float(item.get("product_price", 0))

        confirmed = int(item.get("confirmed", 0))
        status = int(item.get("status", 0))
        email = item.get("utm_source", "")

        if confirmed == 1:

            approved_commission += commission
            approved_sales += sales
            
            _, a_ratio, _ = get_user_ratios(email)
            approved_admin_profit += commission * a_ratio

            approved_count += 1
            order_status = 1

        elif status == 2:

            rejected_commission += commission
            rejected_sales += sales

            reject_count += 1
            order_status = 2

        else:

            pending_commission += commission
            pending_sales += sales

            pending_count += 1
            order_status = 0

        result.append({
            "order_id": item.get("order_id"),
            "order_time": item.get("sales_time"),
            "campaign_name": item.get("campaign_id"),
            "sales_amount": sales,
            "pub_commission": commission,
            "order_status": order_status,
            "utm_source": item.get("utm_source", "")
        })
        
    result.sort(key=lambda x: x["order_time"], reverse=True)
    return {
        "success": True,
        "summary": {

            "conversions": total_orders,

            "approved_commission": round(approved_commission),

            "pending_commission": round(pending_commission),

            "rejected_commission": round(rejected_commission),

            "approved_sales": round(approved_sales),

            "pending_sales": round(pending_sales),

            "rejected_sales": round(rejected_sales),
            "net_profit": round(approved_admin_profit),

            "generated_links": firebase["generated_links"],

            "users": firebase["users"],

            "logs": firebase["logs"]

        },
        "analytics": {**analytics, "order_status": {"approved": approved_count, "pending": pending_count, "rejected": reject_count}},
        "orders": result[:30]
    }

# ==========================================
# ENDPOINTS: RÚT TIỀN (WITHDRAWALS)
# ==========================================
@app.post("/api/withdrawals")
@limiter.limit("30/minute") 
async def create_withdrawal(request: Request, body: WithdrawalRequest):
    # ĐÃ FIX LỖI: Truyền object request vào hàm get_user_wallet
    wallet = get_user_wallet(body.user_email, request)
    
    pending_request = db.collection("withdrawals")\
    .where("user_email","==",body.user_email)\
    .where("status","==","pending")\
    .stream()
    
    pending_amount = sum(w.to_dict()["amount"] for w in pending_request)
    
    if body.amount > wallet["balance"] - pending_amount:
        raise HTTPException(status_code=400, detail="Số dư khả dụng không đủ.")
        
    if body.amount < 100000:
        raise HTTPException(status_code=400, detail="Rút tối thiểu 100.000đ")
        
    db.collection("withdrawals").add({
        "user_email": body.user_email,
        "amount": body.amount,
        "bank_info": body.bank_info,
        "status": "pending",
        "balance_at_request": wallet["balance"],
        "created_at": firestore.SERVER_TIMESTAMP
    })
    return {"success": True, "message": "Yêu cầu rút tiền đã được gửi thành công!"}

@app.get("/api/admin/withdrawals")
@limiter.limit("30/minute")
def get_withdrawals(request: Request):
    """Admin lấy danh sách yêu cầu rút tiền"""
    verify_admin(request)
    docs = db.collection("withdrawals").order_by("created_at", direction=firestore.Query.DESCENDING).stream()
    
    result = []
    for doc in docs:
        data = doc.to_dict()
        created_time = ""
        if "created_at" in data and data["created_at"]:
            try:
                created_time = data["created_at"].strftime("%d/%m/%Y %H:%M")
            except:
                created_time = str(data["created_at"])
        result.append({
            "id": doc.id,
            "email": data.get("user_email", "N/A"),
            "amount": data.get("amount", 0),
            "bank": data.get("bank_info", "N/A"),
            "status": data.get("status", "pending"),
            "date": created_time
        })
    return {"success": True, "data": result}

@app.post("/api/admin/withdrawals/update")
@limiter.limit("30/minute")
def update_withdrawal(request: Request, body: WithdrawalUpdate):
    """Admin cập nhật trạng thái yêu cầu rút tiền"""
    verify_admin(request)
    doc_ref = db.collection("withdrawals").document(body.request_id)
    
    if not doc_ref.get().exists:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu rút tiền này")
        
    doc_ref.update({
        "status": body.status,
        "updated_at": datetime.now()
    })
    return {"success": True, "message": "Cập nhật trạng thái thành công"}

# ==========================================
# ENDPOINT: USER LẤY LỊCH SỬ RÚT TIỀN 
# ==========================================
@app.get("/api/user/withdrawals/history")
@limiter.limit("30/minute")
def get_user_withdrawals_history(email: str, request: Request):
    # Lấy dữ liệu và tự sắp xếp để tránh lỗi Index của Firebase
    docs = db.collection("withdrawals").where("user_email", "==", email).stream()
    
    result = []
    for doc in docs:
        data = doc.to_dict()
        created_time = ""
        time_val = 0
        
        if "created_at" in data and data["created_at"]:
            try:
                if hasattr(data["created_at"], "timestamp"):
                    time_val = data["created_at"].timestamp()
                    created_time = data["created_at"].strftime("%d/%m/%Y %H:%M")
                elif isinstance(data["created_at"], str):
                    dt = datetime.fromisoformat(data["created_at"][:19])
                    time_val = dt.timestamp()
                    created_time = dt.strftime("%d/%m/%Y %H:%M")
            except:
                created_time = str(data["created_at"])
                
        result.append({
            "id": doc.id,
            "amount": data.get("amount", 0),
            "bank": data.get("bank_info", "N/A"),
            "status": data.get("status", "pending"),
            "date": created_time,
            "time_val": time_val
        })
        
    # Sắp xếp danh sách dựa vào timestamp (mới nhất lên đầu)
    result.sort(key=lambda x: x["time_val"], reverse=True)
    
    # Dọn dẹp dữ liệu thừa
    for r in result:
        del r["time_val"]
        
    return {"success": True, "data": result}

# Lấy lịch sử đơn hàng của 1 user
@app.get("/api/user/history")
@limiter.limit("30/minute")
def get_user_history(email:str, request: Request):
    orders = db.collection("orders")\
        .where("utm_source","==",email)\
        .stream()
    result = []
    for doc in orders:
        order = doc.to_dict()
        if order.get("utm_source")!=email:
            continue
        u_ratio, _, _ = get_user_ratios(email)
        cashback = float(order["reward"]) * u_ratio
        if order.get("confirmed")==1:
            status="approved"

        elif order.get("status")==2:
            status="rejected"

        else:
            status="pending"
        result.append({

            "order_id": order["order_id"],

            "merchant": order.get("campaign_id"),

            "amount": order.get("product_price",0),

            "cashback": round(cashback),

            "status": status,

            "time": order.get("sales_time")

        })
    result.sort(key=lambda x:x["time"], reverse=True)
    return {"success":True, "orders":result}

# ==========================================
# ENDPOINT: QUẢN LÝ NGƯỜI DÙNG (ADMIN)
# ==========================================
@app.get("/api/admin/users")
@limiter.limit("30/minute")
def get_admin_users(request: Request, start_date: str = None, end_date: str = None):
    """Admin lấy danh sách chi tiết hành vi người dùng có lọc theo ngày chuẩn xác"""
    verify_admin(request)
    
    conversions = db.collection("conversions").stream()
    user_data = defaultdict(lambda: {"email": "", "total_links": 0, "recent_links": []})
    
    # Ép kiểu thẳng về chuẩn Date (YYYY-MM-DD) để so sánh, bỏ qua timezone
    start_d = None
    end_d = None
    if start_date:
        try: start_d = datetime.strptime(start_date, "%Y-%m-%d").date()
        except: pass
    if end_date:
        try: end_d = datetime.strptime(end_date, "%Y-%m-%d").date()
        except: pass

    for doc in conversions:
        data = doc.to_dict()
        email = data.get("user_email")
        if not email:
            continue
            
        created_at = data.get("created_at")
        time_str = "N/A"
        time_val = 0
        doc_d = None
        
        if created_at:
            try:
                # Cộng 7 tiếng để về chuẩn giờ Việt Nam
                if hasattr(created_at, "timestamp"):
                    time_val = created_at.timestamp()
                    dt_vn = datetime.fromtimestamp(time_val, tz=timezone.utc) + timedelta(hours=7)
                    doc_d = dt_vn.date()
                    time_str = dt_vn.strftime("%d/%m/%Y %H:%M")
                elif isinstance(created_at, str):
                    dt_obj = datetime.fromisoformat(created_at[:19])
                    doc_d = dt_obj.date()
                    time_val = dt_obj.timestamp()
                    time_str = dt_obj.strftime("%d/%m/%Y %H:%M")
            except Exception:
                pass

        # NẾU CÓ BỘ LỌC NGÀY NHƯNG DỮ LIỆU KHÔNG PARSE ĐƯỢC NGÀY -> BỎ QUA LUÔN
        if (start_d or end_d) and not doc_d:
            continue
            
        # NẾU KHÔNG NẰM TRONG KHOẢNG NGÀY -> BỎ QUA
        if doc_d:
            if start_d and doc_d < start_d:
                continue
            if end_d and doc_d > end_d:
                continue

        user_data[email]["email"] = email
        user_data[email]["total_links"] += 1
        
        user_data[email]["recent_links"].append({
            "product_name": data.get("product_name", "N/A"),
            "platform": data.get("platform", "N/A"),
            "short_link": data.get("short_link", "N/A"),
            "time_str": time_str,
            "time_val": time_val
        })
        
    result = []
    for email, info in user_data.items():
        info["recent_links"].sort(key=lambda x: x["time_val"], reverse=True)
        recent = info["recent_links"]
        # CHỈ THÊM USER NÀO CÓ TẠO LINK TRONG KHOẢNG THỜI GIAN ĐÃ LỌC
        if info["total_links"] > 0:
            result.append({
                "email": email,
                "total_links": info["total_links"],
                "recent_links": recent
            })
            
    result.sort(key=lambda x: x["total_links"], reverse=True)
    return {"success": True, "data": result}

@app.get("/api/leaderboard")
def leaderboard():

    orders = db.collection("orders").stream()

    ranking = defaultdict(float)

    for doc in orders:
        order = doc.to_dict()

        email = order.get("utm_source")
        if not email:
            continue

        if int(order.get("confirmed", 0)) != 1:
            continue

        reward = float(order.get("reward", 0))
        u_ratio, _, _ = get_user_ratios(email)
        ranking[email] += reward * u_ratio

    result = []

    for email, total in ranking.items():

        user_doc = db.collection("users").document(email).get()

        avatar = ""
        name = email.split("@")[0]

        if user_doc.exists:
            user_data = user_doc.to_dict()

            avatar = user_data.get("photoURL", "") or user_data.get("avatar", "")

            if user_data.get("displayName"):
                name = user_data["displayName"]

        result.append({
            "email": email,
            "name": name,
            "avatar": avatar,
            "cashback": round(total)
        })

    result.sort(
        key=lambda x: x["cashback"],
        reverse=True
    )

    return {
        "success": True,
        "data": result[:3]
    }

@app.post("/api/admin/sync-users")
@limiter.limit("5/minute")
def sync_users(request: Request):
    verify_admin(request)

    page = firebase_auth.list_users()

    total = 0

    while page:
        for user in page.users:
            email = user.email
            if not email:
                continue

            db.collection("users").document(email).set(
                {
                    "uid": user.uid,
                    "email": email,
                    "displayName": user.display_name or email.split("@")[0],
                    "photoURL": user.photo_url or "",
                    "phoneNumber": user.phone_number or "",
                    "disabled": user.disabled,
                    "createdAt": user.user_metadata.creation_timestamp,
                    "lastSignIn": user.user_metadata.last_sign_in_timestamp,
                    "provider": (
                        user.provider_data[0].provider_id
                        if user.provider_data
                        else ""
                    ),
                },
                merge=True,
            )

            total += 1

        page = page.get_next_page()

    return {
        "success": True,
        "synced": total
    }

@app.post("/api/admin/import-shopee-report")
@limiter.limit("5/minute")
async def import_shopee_report(request: Request, file: UploadFile = File(...)):
    verify_admin(request)
    
    import openpyxl
    from io import BytesIO
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), read_only=True)
        sheet = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không thể đọc file Excel: {str(e)}")
    
    header_row = None
    headers = []
    
    # Quét tối đa 10 dòng đầu để tìm dòng tiêu đề
    for r_idx in range(1, 11):
        try:
            row_vals = [str(sheet.cell(row=r_idx, column=c_idx).value or "").strip() for c_idx in range(1, sheet.max_column + 1)]
            keywords = ["mã đơn hàng", "order id", "đơn hàng", "sub_id", "sub id", "hoa hồng", "commission"]
            if any(any(kw in val.lower() for kw in keywords) for val in row_vals):
                header_row = r_idx
                headers = row_vals
                break
        except Exception:
            continue
            
    if not header_row:
        raise HTTPException(status_code=400, detail="Không tìm thấy dòng tiêu đề hợp lệ trong file Excel. Vui lòng kiểm tra lại file báo cáo của Shopee.")
        
    col_map = {}
    for idx, name in enumerate(headers):
        name_lower = name.lower()
        if "mã đơn hàng" in name_lower or "order id" in name_lower or "order_id" in name_lower:
            col_map["order_id"] = idx
        elif "trạng thái" in name_lower or "status" in name_lower:
            col_map["status"] = idx
        elif "giá trị đơn" in name_lower or "order value" in name_lower or "doanh số" in name_lower or "giá trị sản phẩm" in name_lower or "product price" in name_lower or "price" in name_lower:
            col_map["product_price"] = idx
        elif "hoa hồng" in name_lower or "commission" in name_lower or "reward" in name_lower:
            col_map["reward"] = idx
        elif "sub_id" in name_lower or "sub id" in name_lower or "sub-id" in name_lower or "utm_content" in name_lower:
            col_map["sub_id"] = idx
        elif "thời gian" in name_lower or "time" in name_lower or "sales_time" in name_lower:
            col_map["sales_time"] = idx
        elif "mã lượt click" in name_lower or "click id" in name_lower or "transaction_id" in name_lower:
            col_map["transaction_id"] = idx

    required_cols = ["order_id", "status", "reward"]
    missing = [c for c in required_cols if c not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"File Excel thiếu các cột bắt buộc: {', '.join(missing)}")

    success_count = 0
    skipped_count = 0
    
    # Đọc dữ liệu từ các dòng tiếp theo
    for r_idx in range(header_row + 1, sheet.max_row + 1):
        try:
            row_vals = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, len(headers) + 1)]
            if not row_vals or all(val is None for val in row_vals):
                continue
                
            def get_val(col_name, default=None):
                idx = col_map.get(col_name)
                if idx is not None and idx < len(row_vals):
                    return row_vals[idx]
                return default

            raw_order_id = get_val("order_id")
            if not raw_order_id:
                skipped_count += 1
                continue
                
            order_id = str(raw_order_id).strip()
            raw_status = get_val("status", "")
            status_str = str(raw_status).strip().lower()
            
            confirmed = 0
            status_code = 0
            
            if any(x in status_str for x in ["hoàn thành", "thành công", "completed", "đã hoàn thành"]):
                confirmed = 1
                status_code = 1
            elif any(x in status_str for x in ["hủy", "cancelled", "không thành công", "đã hủy"]):
                confirmed = 0
                status_code = 2
            else:
                confirmed = 0
                status_code = 0
                
            def parse_float(val):
                if val is None:
                    return 0.0
                try:
                    s = str(val).replace(",", "").replace("đ", "").replace("VND", "").strip()
                    return float(s)
                except:
                    return 0.0

            reward = parse_float(get_val("reward"))
            product_price = parse_float(get_val("product_price"))
            
            raw_sub_id = get_val("sub_id", "")
            sub_id = str(raw_sub_id).strip()
            
            email = ""
            if "hangthocashback-" in sub_id:
                sanitized_email = sub_id.split("hangthocashback-")[1]
                if "_at_" in sanitized_email:
                    parts = sanitized_email.split("_at_")
                    prefix = parts[0]
                    domain = parts[1].replace("_", ".")
                    email = f"{prefix}@{domain}"
            else:
                if "@" in sub_id:
                    email = sub_id

            raw_tx_id = get_val("transaction_id")
            if raw_tx_id:
                transaction_id = str(raw_tx_id).strip()
            else:
                transaction_id = f"shopee_{order_id}"

            sales_time = get_val("sales_time")
            if sales_time:
                if hasattr(sales_time, "strftime"):
                    sales_time_str = sales_time.strftime("%Y-%m-%dT%H:%M:%S")
                else:
                    sales_time_str = str(sales_time).strip().replace(" ", "T")
            else:
                sales_time_str = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

            db.collection("orders").document(transaction_id).set({
                "transaction_id": transaction_id,
                "order_id": order_id,
                "campaign_id": "Shopee",
                "product_id": "",
                "quantity": 1,
                "product_price": product_price,
                "reward": reward,
                "sales_time": sales_time_str,
                "status": status_code,
                "confirmed": confirmed,
                "utm_source": email,
                "utm_content": sub_id,
                "utm_medium": "shopee",
                "created_at": firestore.SERVER_TIMESTAMP
            }, merge=True)
            
            success_count += 1
        except Exception as e:
            print(f"Lỗi khi đọc dòng {r_idx}: {e}")
            skipped_count += 1
            
    return {
        "success": True,
        "message": f"Đã nhập thành công {success_count} đơn hàng Shopee. Bỏ qua {skipped_count} dòng lỗi."
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)